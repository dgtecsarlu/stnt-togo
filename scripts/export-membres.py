#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
STNT — Extraction des membres en Excel + CSV
============================================
Génère des listes de membres à jour depuis Supabase, pour :
  - les VOTES (corps électoral = membres validés)
  - les RAPPORTS MENSUELS et ANNUELS (avec statistiques)
  - un export COMPLET

Sortie : un .xlsx (avec onglet Statistiques) ET un .csv (UTF-8 BOM,
accents lisibles dans Excel), dans stnt-togo/exports/.

Identifiants requis (lecture seule conseillée, ici clé service car la
table membres est protégée par RLS) — via variables d'environnement ou
le .env à la racine du workspace :
    SUPABASE_URL=https://xxxx.supabase.co
    SUPABASE_SERVICE_ROLE_KEY=eyJ...
La clé service NE DOIT JAMAIS être commitée (déjà couverte par .gitignore).

Exemples :
    python export-membres.py --scope votes
    python export-membres.py --scope mois  --periode 2026-06
    python export-membres.py --scope annee --periode 2026
    python export-membres.py --scope tous
"""
import argparse
import csv
import json
import os
import sys
import urllib.request
import urllib.parse
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Colonnes exportées (ordre d'affichage) et leur libellé
COLS = [
    ("nom_complet", "Nom et prénoms"),
    ("email", "Email"),
    ("telephone", "Téléphone"),
    ("region", "Région"),
    ("ville", "Ville"),
    ("metier", "Métier"),
    ("type_adhesion", "Type"),
    ("statut_cotisation", "Cotisation"),
    ("statut_validation", "Validation"),
    ("date_adhesion", "Date d'adhésion"),
]

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))


def load_env():
    """Charge SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY depuis l'env ou le .env racine."""
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    env_path = os.path.join(ROOT, ".env")
    if (not url or not key) and os.path.exists(env_path):
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                k, v = line.split("=", 1)
                k, v = k.strip(), v.strip().strip('"').strip("'")
                if k == "SUPABASE_URL" and not url:
                    url = v
                if k == "SUPABASE_SERVICE_ROLE_KEY" and not key:
                    key = v
    return url, key


def fetch_membres(url, key, filtre):
    """Lit la table membres via PostgREST (la clé service contourne la RLS)."""
    params = {"select": "*", "order": "nom_complet.asc"}
    params.update(filtre)
    endpoint = url.rstrip("/") + "/rest/v1/membres?" + urllib.parse.urlencode(params)
    req = urllib.request.Request(endpoint, headers={
        "apikey": key,
        "Authorization": "Bearer " + key,
        "Accept": "application/json",
    })
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read().decode("utf-8"))


def periode_filtre(scope, periode):
    """Construit le filtre date PostgREST selon le scope."""
    if scope == "votes":
        return {"statut_validation": "eq.validee"}, "corps-electoral"
    if scope == "tous":
        return {}, "tous"
    if scope == "mois":
        # periode = YYYY-MM
        an, mois = periode.split("-")
        an, mois = int(an), int(mois)
        debut = f"{an:04d}-{mois:02d}-01"
        fin = f"{an+1:04d}-01-01" if mois == 12 else f"{an:04d}-{mois+1:02d}-01"
        return ({"date_adhesion": f"gte.{debut}", "and": f"(date_adhesion.lt.{fin})"},
                f"mensuel-{an:04d}-{mois:02d}")
    if scope == "annee":
        an = int(periode)
        return ({"date_adhesion": f"gte.{an:04d}-01-01", "and": f"(date_adhesion.lt.{an+1:04d}-01-01)"},
                f"annuel-{an:04d}")
    raise ValueError("scope inconnu")


def stats(rows):
    """Comptages pour les rapports."""
    def compte(champ):
        d = {}
        for r in rows:
            v = r.get(champ) or "(non renseigné)"
            d[v] = d.get(v, 0) + 1
        return d
    return {
        "Total membres": len(rows),
        "Par validation": compte("statut_validation"),
        "Par cotisation": compte("statut_cotisation"),
        "Par type": compte("type_adhesion"),
        "Par région": compte("region"),
        "Avec email": sum(1 for r in rows if r.get("email")),
    }


def ecrire_excel(rows, st, chemin, titre):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Membres"
    blue = PatternFill("solid", fgColor="081826")
    head = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="B0BEC5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    libelles = [lib for _, lib in COLS]
    ws.append(libelles)
    for c in range(1, len(libelles) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = blue; cell.font = head; cell.alignment = center; cell.border = border

    for r in rows:
        ws.append([fmt(r.get(champ)) for champ, _ in COLS])

    widths = [30, 28, 16, 12, 14, 18, 10, 12, 12, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for ri in range(2, ws.max_row + 1):
        for ci in range(1, len(libelles) + 1):
            ws.cell(row=ri, column=ci).border = border
    ws.freeze_panes = "A2"
    if ws.max_row >= 1:
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(libelles))}{ws.max_row}"

    # Onglet statistiques (utile pour les rapports)
    wss = wb.create_sheet("Statistiques")
    wss["A1"] = titre; wss["A1"].font = Font(bold=True, size=13, color="081826")
    wss["A2"] = "Généré le " + datetime.now().strftime("%Y-%m-%d %H:%M")
    ligne = 4
    for cle, val in st.items():
        wss.cell(row=ligne, column=1, value=cle).font = Font(bold=True)
        if isinstance(val, dict):
            for k, v in sorted(val.items(), key=lambda x: -x[1]):
                wss.cell(row=ligne, column=2, value=str(k))
                wss.cell(row=ligne, column=3, value=v)
                ligne += 1
        else:
            wss.cell(row=ligne, column=3, value=val)
            ligne += 1
        ligne += 1
    wss.column_dimensions["A"].width = 20
    wss.column_dimensions["B"].width = 22
    wb.save(chemin)


def ecrire_csv(rows, chemin):
    # utf-8-sig : BOM pour que Excel affiche correctement les accents
    with open(chemin, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow([lib for _, lib in COLS])
        for r in rows:
            w.writerow([fmt(r.get(champ)) for champ, _ in COLS])


def fmt(v):
    if v is None:
        return ""
    if isinstance(v, str) and "T" in v and v[:4].isdigit():
        return v.split("T")[0]  # date seule
    return v


def main():
    p = argparse.ArgumentParser(description="Export des membres STNT (Excel + CSV)")
    p.add_argument("--scope", required=True, choices=["votes", "mois", "annee", "tous"])
    p.add_argument("--periode", help="YYYY-MM pour --scope mois, YYYY pour --scope annee")
    args = p.parse_args()

    if args.scope in ("mois", "annee") and not args.periode:
        sys.exit("Erreur : --periode est requis pour --scope " + args.scope)

    url, key = load_env()
    if not url or not key:
        sys.exit("Erreur : SUPABASE_URL et SUPABASE_SERVICE_ROLE_KEY introuvables "
                 "(variables d'environnement ou .env racine).")

    filtre, suffixe = periode_filtre(args.scope, args.periode)
    try:
        rows = fetch_membres(url, key, filtre)
    except Exception as e:
        sys.exit("Erreur de lecture Supabase : " + str(e))

    out_dir = os.path.join(os.path.dirname(__file__), "..", "exports")
    os.makedirs(out_dir, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d")
    base = os.path.join(out_dir, f"membres-{suffixe}-{stamp}")

    st = stats(rows)
    titre = f"STNT — Export membres ({suffixe})"
    ecrire_excel(rows, st, base + ".xlsx", titre)
    ecrire_csv(rows, base + ".csv")

    print(f"OK : {len(rows)} membre(s) exporté(s)")
    print(" -", base + ".xlsx")
    print(" -", base + ".csv")
    print(" - Total avec email :", st["Avec email"])


if __name__ == "__main__":
    main()
